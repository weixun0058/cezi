import argparse
import csv
import json
import os
import sqlite3
from pathlib import Path

from opencc import OpenCC
from openpyxl import load_workbook

GUA_FIELDS = {
    "签号": "sign_number",
    "吉凶": "fortune",
    "卦属": "gua_type",
    "签文": "sign_text",
    "解签一": "interpretation1",
    "事业": "career",
    "财运": "wealth",
    "情感": "love",
    "健康": "health",
    "学业": "study",
    "泛泛": "general",
}
GUA_COLUMNS = tuple(GUA_FIELDS.values())
INTERPRETATION_FIELDS = (
    "interpretation1",
    "career",
    "wealth",
    "love",
    "health",
    "study",
    "general",
)


def _read_hanzi(source, supplement_source=None):
    connection = sqlite3.connect(source)
    try:
        rows = connection.execute("""
            SELECT 汉字, 简体字总笔画, 繁体字总笔画, 康熙字典笔画
            FROM hanzi
            WHERE 汉字 IS NOT NULL AND 汉字 != ''
            ORDER BY 汉字, ID
            """).fetchall()
        unique = {}
        for row in rows:
            unique.setdefault(row[0], row)
        if supplement_source:
            supplement_path = Path(supplement_source)
            if supplement_path.exists():
                with supplement_path.open(encoding="utf-8-sig", newline="") as file:
                    for item in csv.DictReader(file):
                        character = (item.get("character") or "").strip()
                        try:
                            kangxi_strokes = int(item.get("kangxi_strokes") or 0)
                        except ValueError:
                            continue
                        if character and kangxi_strokes > 0:
                            unique.setdefault(character, (character, 0, 0, kangxi_strokes))
        return sorted(unique.values(), key=lambda item: item[0])
    finally:
        connection.close()


def _read_gua(source, reinterpretations_source=None):
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        records = []
        for values in rows:
            row = dict(zip(headers, values, strict=False))
            try:
                sign_number = int(row.get("签号"))
            except (TypeError, ValueError):
                continue
            records.append(
                tuple(
                    sign_number if field == "sign_number" else row.get(source_name) or ""
                    for source_name, field in GUA_FIELDS.items()
                )
            )
        by_number = {record[0]: list(record) for record in records}
        if reinterpretations_source:
            reinterpretations_path = Path(reinterpretations_source)
            if reinterpretations_path.exists():
                with reinterpretations_path.open(encoding="utf-8") as file:
                    for item in json.load(file):
                        record = by_number.get(item.get("sign_number"))
                        if record is None:
                            continue
                        for field in INTERPRETATION_FIELDS:
                            record[GUA_COLUMNS.index(field)] = item.get(field) or ""
        return [tuple(by_number[number]) for number in sorted(by_number)]
    finally:
        workbook.close()


def _read_pzbj(source):
    with open(source, encoding="utf-8") as file:
        return sorted(json.load(file).items())


def _build_gua_hant(gua, source=None):
    converter = OpenCC("s2t")
    by_number = {
        row[0]: [row[0], *[converter.convert(row[index]) for index in range(1, len(row))]]
        for row in gua
    }
    if source:
        path = Path(source)
        if path.exists():
            with path.open(encoding="utf-8") as file:
                for item in json.load(file):
                    record = by_number.get(item.get("sign_number"))
                    if record is None:
                        continue
                    for field in GUA_COLUMNS[1:]:
                        if field in item:
                            record[GUA_COLUMNS.index(field)] = item[field] or ""
    return [tuple(by_number[number]) for number in sorted(by_number)]


def build_reference_database(
    output,
    hanzi_source,
    gua_source,
    pzbj_source,
    hanzi_supplement_source=None,
    gua_reinterpretations_source=None,
    gua_hant_source=None,
):
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.name}.tmp")
    temporary.unlink(missing_ok=True)

    if hanzi_supplement_source is None:
        hanzi_supplement_source = Path(hanzi_source).with_name("hanzi_strokes_zdic.csv")
    if gua_reinterpretations_source is None:
        gua_reinterpretations_source = (
            Path(gua_source).parents[1] / "content" / "oracle_signs_reinterpreted.json"
        )
    if gua_hant_source is None:
        gua_hant_source = (
            Path(gua_source).parents[1] / "content" / "oracle_signs_reinterpreted_hant.json"
        )
    hanzi = _read_hanzi(hanzi_source, hanzi_supplement_source)
    gua = _read_gua(gua_source, gua_reinterpretations_source)
    pzbj = _read_pzbj(pzbj_source)
    connection = sqlite3.connect(temporary)
    try:
        connection.executescript("""
            PRAGMA page_size=4096;
            PRAGMA journal_mode=DELETE;
            PRAGMA synchronous=FULL;
            CREATE TABLE hanzi (
                character TEXT PRIMARY KEY,
                simplified_strokes INTEGER,
                traditional_strokes INTEGER,
                kangxi_strokes INTEGER
            ) WITHOUT ROWID;
            CREATE TABLE gua (
                sign_number INTEGER PRIMARY KEY CHECK(sign_number BETWEEN 1 AND 384),
                fortune TEXT NOT NULL,
                gua_type TEXT NOT NULL,
                sign_text TEXT NOT NULL,
                interpretation1 TEXT NOT NULL,
                career TEXT NOT NULL,
                wealth TEXT NOT NULL,
                love TEXT NOT NULL,
                health TEXT NOT NULL,
                study TEXT NOT NULL,
                general TEXT NOT NULL
            );
            CREATE TABLE pzbj (
                text TEXT PRIMARY KEY,
                explanation TEXT NOT NULL
            ) WITHOUT ROWID;
            CREATE TABLE gua_hant (
                sign_number INTEGER PRIMARY KEY CHECK(sign_number BETWEEN 1 AND 384),
                fortune TEXT NOT NULL,
                gua_type TEXT NOT NULL,
                sign_text TEXT NOT NULL,
                interpretation1 TEXT NOT NULL,
                career TEXT NOT NULL,
                wealth TEXT NOT NULL,
                love TEXT NOT NULL,
                health TEXT NOT NULL,
                study TEXT NOT NULL,
                general TEXT NOT NULL
            );
            CREATE TABLE pzbj_hant (
                text TEXT PRIMARY KEY,
                explanation TEXT NOT NULL
            ) WITHOUT ROWID;
            """)
        connection.executemany("INSERT INTO hanzi VALUES (?, ?, ?, ?)", hanzi)
        connection.executemany("INSERT INTO gua VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", gua)
        connection.executemany("INSERT INTO pzbj VALUES (?, ?)", pzbj)

        # 繁体签文优先使用运行时同一份权威 JSON，缺失字段才回退 OpenCC。
        converter = OpenCC("s2t")
        gua_hant = _build_gua_hant(gua, gua_hant_source)
        pzbj_hant = [
            (converter.convert(text), converter.convert(explanation)) for text, explanation in pzbj
        ]
        connection.executemany(
            "INSERT INTO gua_hant VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", gua_hant
        )
        connection.executemany("INSERT INTO pzbj_hant VALUES (?, ?)", pzbj_hant)

        connection.commit()
        connection.execute("VACUUM")
    finally:
        connection.close()

    os.replace(temporary, output)
    return {
        "hanzi": len(hanzi),
        "gua": len(gua),
        "pzbj": len(pzbj),
        "gua_hant": len(gua_hant),
        "pzbj_hant": len(pzbj_hant),
    }


def main():
    parser = argparse.ArgumentParser(description="Build the production reference database")
    parser.add_argument("--output", default="data/reference/reference.db")
    parser.add_argument("--hanzi", default="data/reference/kanxi_dict.db")
    parser.add_argument(
        "--hanzi-supplement",
        default="data/reference/hanzi_strokes_zdic.csv",
        help="汉典补充笔画权威源",
    )
    parser.add_argument("--gua", default="data/reference/zhugeshenshuan_jq.xlsx")
    parser.add_argument(
        "--gua-reinterpretations",
        default="data/content/oracle_signs_reinterpreted.json",
        help="384 签新版解签权威源",
    )
    parser.add_argument(
        "--gua-hant",
        default="data/content/oracle_signs_reinterpreted_hant.json",
        help="384 签繁体权威源",
    )
    parser.add_argument("--pzbj", default="data/reference/pzbj.json")
    arguments = parser.parse_args()
    counts = build_reference_database(
        arguments.output,
        arguments.hanzi,
        arguments.gua,
        arguments.pzbj,
        arguments.hanzi_supplement,
        arguments.gua_reinterpretations,
        arguments.gua_hant,
    )
    print("Built reference database: " + ", ".join(f"{k}={v}" for k, v in counts.items()))


if __name__ == "__main__":
    main()
