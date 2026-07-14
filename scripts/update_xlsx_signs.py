"""用权威简体签文 CSV 更新 xlsx 签文列（第 4 列）。

读取 oracle_signs_authoritative_sc.csv，
替换 zhugeshenshuan_jq.xlsx 中第 4 列（签文）的内容，
保留其他列不变。

备份策略：修改前自动备份为 .bak。
"""

import csv
import shutil
from pathlib import Path

import openpyxl

SC_CSV = Path("data/reference/oracle_signs_authoritative_sc.csv")
XLSX_PATH = Path("data/reference/zhugeshenshuan_jq.xlsx")
BACKUP_PATH = XLSX_PATH.with_suffix(".xlsx.bak")

SIGN_TEXT_COL = 4  # 第 4 列（1-based）


def load_authoritative_sc():
    """从权威简体 CSV 读取签文。"""
    signs = {}
    with open(SC_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            signs[int(row["sign_number"])] = row["sign_text"]
    return signs


def update_xlsx(sc_signs):
    """更新 xlsx 签文列，返回修改记录。"""
    shutil.copy2(XLSX_PATH, BACKUP_PATH)
    print(f"备份: {BACKUP_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Sheet1"]

    changes = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        sn_cell = row[0].value
        if sn_cell is None:
            continue
        sn = int(sn_cell)
        if sn not in sc_signs:
            continue

        old_text = row[SIGN_TEXT_COL - 1].value or ""
        new_text = sc_signs[sn]
        if old_text != new_text:
            changes.append((sn, old_text, new_text))
            ws.cell(row=row_idx, column=SIGN_TEXT_COL, value=new_text)

    wb.save(XLSX_PATH)
    wb.close()
    return changes


def main():
    sc_signs = load_authoritative_sc()
    print(f"权威简体签数: {len(sc_signs)}")

    changes = update_xlsx(sc_signs)
    print(f"修改签数: {len(changes)}")
    print("\n=== 修改详情 ===")
    for sn, old, new in changes:
        print(f"\n第 {sn} 签:")
        print(f"  旧: {old}")
        print(f"  新: {new}")


if __name__ == "__main__":
    main()
