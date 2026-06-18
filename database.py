import logging
import sqlite3
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

LOGGER = logging.getLogger(__name__)
GUA_FIELDS = {
    "签号": "sign_number",
    "吉凶": "fortune",
    "卦属": "gua_type",
    "签文": "sign_text",
    "解签一": "interpretation1",
    "事业": "career",
    "财运": "wealth",
    "情感": "love",
    "健康": "health",
    "学业": "study",
    "泛泛": "general",
}


class Database:
    def __init__(
        self, hanzi_db="database/kanxi_dict.db", gua_data_path="database/zhugeshenshuan_jq.xlsx"
    ):
        self.hanzi_db = Path(hanzi_db)
        self.gua_data_path = Path(gua_data_path)
        self._stroke_failure_cache = {}
        self._stroke_failure_ttl = 300
        self.gua_index = self._load_gua_index()

    def _load_gua_index(self):
        workbook = load_workbook(self.gua_data_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        index = {}
        for values in rows:
            row = dict(zip(headers, values, strict=False))
            try:
                sign_number = int(row.get("签号"))
            except (TypeError, ValueError):
                continue
            index[sign_number] = {
                target: row.get(source) if row.get(source) is not None else ""
                for source, target in GUA_FIELDS.items()
            }
            index[sign_number]["sign_number"] = sign_number
        workbook.close()
        LOGGER.info("Loaded %d divination records", len(index))
        return index

    def get_gua_info(self, sign_number):
        try:
            return self.gua_index.get(int(sign_number))
        except (TypeError, ValueError):
            return None

    def get_stroke_count_by_hd(self, character):
        """Return a remote stroke count, or None when the lookup is inconclusive."""
        cached_until = self._stroke_failure_cache.get(character, 0)
        if cached_until > time.monotonic():
            return None

        try:
            response = requests.get(
                f"https://www.zdic.net/hans/{character}",
                headers={"User-Agent": "ZhugeshenSuan/4.1 (+stroke lookup)"},
                timeout=(3.05, 5),
            )
            response.raise_for_status()
            response.encoding = "utf-8"
            soup = BeautifulSoup(response.text, "html.parser")

            kx_path = soup.select("div.kxzd span.res_d a")
            if kx_path:
                return int(kx_path[-1].get_text(strip=True))

            for node in soup.select("td.z_bs2 p"):
                if "总笔画" in node.get_text():
                    return int(node.get_text().split()[-1])
        except (requests.RequestException, TypeError, ValueError) as exc:
            LOGGER.warning("Remote stroke lookup failed for one character: %s", type(exc).__name__)

        self._stroke_failure_cache[character] = time.monotonic() + self._stroke_failure_ttl
        return None

    def get_stroke_count(self, character):
        if not isinstance(character, str) or len(character) != 1:
            return None

        try:
            with sqlite3.connect(self.hanzi_db) as connection:
                row = connection.execute(
                    """
                    SELECT 简体字总笔画, 繁体字总笔画, 康熙字典笔画
                    FROM hanzi
                    WHERE 汉字 = ?
                    ORDER BY ID
                    LIMIT 1
                    """,
                    (character,),
                ).fetchone()
                if row:
                    for value in (row[2], row[1], row[0]):
                        if isinstance(value, int) and value > 0:
                            return value
        except sqlite3.Error:
            LOGGER.exception("Local stroke database lookup failed")
            return None

        stroke_count = self.get_stroke_count_by_hd(character)
        if not stroke_count or stroke_count < 1:
            return None

        try:
            with sqlite3.connect(self.hanzi_db) as connection:
                connection.execute(
                    "INSERT INTO hanzi (汉字, 康熙字典笔画) VALUES (?, ?)",
                    (character, stroke_count),
                )
        except sqlite3.Error:
            LOGGER.exception("Could not cache a remote stroke count")
        return stroke_count
